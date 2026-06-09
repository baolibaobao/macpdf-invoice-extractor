"""Excel export helpers for parsed invoice records."""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any, Iterable, Literal

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from src.invoice_parser import INVOICE_COLUMNS, OPTIONAL_OUTPUT_COLUMNS


ExportMode = Literal["invoice", "items"]

ITEM_COLUMNS = [
    "明细项目名称",
    "规格型号",
    "单位",
    "数量",
    "单价",
    "明细金额",
    "明细税额",
]

AVAILABLE_OUTPUT_COLUMNS = [*INVOICE_COLUMNS, *OPTIONAL_OUTPUT_COLUMNS, "明细项目名称", "明细金额", "明细税额"]


def build_invoice_dataframe(
    records: Iterable[dict[str, Any]],
    mode: ExportMode = "invoice",
    selected_columns: list[str] | None = None,
) -> pd.DataFrame:
    """Build a DataFrame from parsed invoice records.

    ``invoice`` mode exports one row per invoice. ``items`` mode exports one row
    per invoice line item and repeats invoice-level fields on each row.
    """
    columns = _resolve_columns(selected_columns)
    rows: list[dict[str, Any]] = []

    for record in records:
        if record.get("_status") == "skipped":
            continue
        if mode == "items":
            rows.extend(_expand_item_rows(record))
        else:
            rows.append(_invoice_row(record))

    normalized_rows = [_normalize_row(row, columns) for row in rows]
    return pd.DataFrame(normalized_rows, columns=columns)


def export_invoices_to_excel(
    records: Iterable[dict[str, Any]],
    output_path: str | Path,
    mode: ExportMode = "invoice",
    selected_columns: list[str] | None = None,
) -> Path:
    """Export invoice records to an ``.xlsx`` file using openpyxl."""
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    dataframe = build_invoice_dataframe(records, mode=mode, selected_columns=selected_columns)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        dataframe.to_excel(writer, index=False, sheet_name="发票汇总")
        worksheet = writer.sheets["发票汇总"]
        _format_worksheet(worksheet, dataframe)

    return path


def _resolve_columns(selected_columns: list[str] | None) -> list[str]:
    if selected_columns is None:
        return list(INVOICE_COLUMNS)

    valid_columns = set(AVAILABLE_OUTPUT_COLUMNS)
    columns = [column for column in selected_columns if column in valid_columns]
    return columns or list(INVOICE_COLUMNS)


def _invoice_row(record: dict[str, Any]) -> dict[str, Any]:
    row = {column: record.get(column, "") for column in AVAILABLE_OUTPUT_COLUMNS}
    items = record.get("_items") or []
    if items:
        row["项目名称"] = _join_multiline_values(_item_value(item, "项目名称") for item in items)
        row["税率"] = "；".join(dict.fromkeys(_item_value(item, "税率") for item in items if _item_value(item, "税率")))
        for key in ITEM_COLUMNS:
            source_key = _item_source_key(key)
            row[key] = _join_multiline_values(_item_value(item, source_key) for item in items)
    return row


def _expand_item_rows(record: dict[str, Any]) -> list[dict[str, Any]]:
    items = record.get("_items") or []
    if not items:
        return [_invoice_row(record)]
    if len(items) == 1:
        return [_single_item_row(record, items[0])]

    rows: list[dict[str, Any]] = []
    summary_only_columns = {
        "文件路径",
        "文件名称",
        "发票代码",
        "发票号码",
        "发票类型",
        "开票日期",
        "金额合计",
        "税额合计",
        "价税合计",
        "价税合计大写",
        "购买方名称",
        "购买方纳税人识别号",
        "销售方名称",
        "销售方纳税人识别号",
        "项目名称",
        "备注",
    }
    detail_columns = {
        "明细项目名称",
        "规格型号",
        "单位",
        "数量",
        "单价",
        "税率",
        "明细金额",
        "明细税额",
    }
    all_item_names = _join_multiline_values(_item_value(item, "项目名称") for item in items)

    summary_row = {column: record.get(column, "") for column in AVAILABLE_OUTPUT_COLUMNS}
    summary_row["项目名称"] = all_item_names
    for column in detail_columns:
        summary_row[column] = ""
    rows.append(summary_row)

    for item in items:
        row = _single_item_row(record, item)
        for column in summary_only_columns:
            row[column] = ""
        rows.append(row)
    return rows


def _single_item_row(record: dict[str, Any], item: dict[str, Any]) -> dict[str, Any]:
    row = {column: record.get(column, "") for column in AVAILABLE_OUTPUT_COLUMNS}
    row["明细项目名称"] = _item_value(item, "项目名称")
    row["税率"] = _item_value(item, "税率") or row.get("税率", "")
    row["规格型号"] = _item_value(item, "规格型号")
    row["单位"] = _item_value(item, "单位")
    row["数量"] = _item_value(item, "数量")
    row["单价"] = _item_value(item, "单价")
    row["明细金额"] = _item_value(item, "金额")
    row["明细税额"] = _item_value(item, "税额")
    return row


def _item_source_key(output_key: str) -> str:
    if output_key == "明细项目名称":
        return "项目名称"
    if output_key == "明细金额":
        return "金额"
    if output_key == "明细税额":
        return "税额"
    return output_key


def _item_value(item: dict[str, Any], key: str) -> str:
    return str(item.get(key, "") or "").strip()


def _join_multiline_values(values: Iterable[str]) -> str:
    cleaned = [value.strip() for value in values if value and value.strip()]
    return "\n".join(dict.fromkeys(cleaned))


def _normalize_row(row: dict[str, Any], columns: list[str]) -> dict[str, str]:
    return {column: str(row.get(column, "") or "") for column in columns}


def _format_worksheet(worksheet, dataframe: pd.DataFrame) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAD3")
    header_font = Font(bold=True)
    border = Border(
        left=Side(style="thin", color="D9D9D9"),
        right=Side(style="thin", color="D9D9D9"),
        top=Side(style="thin", color="D9D9D9"),
        bottom=Side(style="thin", color="D9D9D9"),
    )

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    worksheet.row_dimensions[1].height = 24

    for row_index, row in enumerate(worksheet.iter_rows(min_row=2), start=2):
        worksheet.row_dimensions[row_index].height = 38
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border

    for index, column_name in enumerate(dataframe.columns, start=1):
        width = _estimate_column_width(column_name, dataframe[column_name].tolist())
        worksheet.column_dimensions[get_column_letter(index)].width = width

    worksheet.freeze_panes = "A2"
    worksheet.auto_filter.ref = worksheet.dimensions


def _estimate_column_width(column_name: str, values: list[Any]) -> int:
    max_len = len(column_name)
    for value in values:
        text = str(value or "")
        max_len = max(max_len, *(len(part) for part in text.splitlines() or [""]))
    return min(max(max_len + 2, 10), 45)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Export parsed invoice JSON records to Excel.")
    parser.add_argument("json_path", help="JSON file containing a list of parsed invoice records")
    parser.add_argument("output_path", help="Output .xlsx path")
    parser.add_argument("--mode", choices=["invoice", "items"], default="invoice")
    parser.add_argument("--columns", nargs="*", help="Optional output columns")
    args = parser.parse_args(argv)

    with Path(args.json_path).open("r", encoding="utf-8") as file:
        records = json.load(file)

    export_invoices_to_excel(records, args.output_path, mode=args.mode, selected_columns=args.columns)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
