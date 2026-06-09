from pathlib import Path

from openpyxl import load_workbook

from src.excel_builder import build_invoice_dataframe, export_invoices_to_excel
from src.invoice_parser import INVOICE_COLUMNS


def _sample_record() -> dict:
    return {
        "文件路径": "D:/invoice/a.pdf",
        "文件名称": "a.pdf",
        "发票代码": "",
        "发票号码": "26317000002019064660",
        "发票类型": "电子发票（普通发票）",
        "开票日期": "2026-06-02",
        "金额合计": "55.66",
        "税额合计": "3.34",
        "税率": "6%",
        "价税合计": "59.00",
        "价税合计大写": "伍拾玖圆整",
        "购买方名称": "深圳星传数字技术有限公司",
        "购买方纳税人识别号": "91440300MADNW7N28X",
        "销售方名称": "中国太平洋财产保险股份有限公司上海分公司",
        "销售方纳税人识别号": "913100007362379322",
        "备注": "保单号:ASHH5202OS26W005JPD0",
        "项目名称": "*保险服务*航空旅客人身\n意外伤害保险（互联网）",
        "_status": "ok",
        "_items": [
            {
                "项目名称": "*保险服务*航空旅客人身\n意外伤害保险（互联网）",
                "规格型号": "**",
                "单位": "单",
                "数量": "1",
                "单价": "55.66",
                "金额": "55.66",
                "税率": "6%",
                "税额": "3.34",
            }
        ],
    }


def test_invoice_mode_uses_standard_16_columns() -> None:
    dataframe = build_invoice_dataframe([_sample_record()])

    assert dataframe.columns.tolist() == INVOICE_COLUMNS
    assert dataframe.iloc[0]["发票号码"] == "26317000002019064660"
    assert "意外伤害保险（互联网）" in dataframe.iloc[0]["项目名称"]


def test_items_mode_expands_each_line_item() -> None:
    record = _sample_record()
    record["_items"].append(
        {
            "项目名称": "*服务*第二项目",
            "规格型号": "",
            "单位": "次",
            "数量": "2",
            "单价": "10.00",
            "金额": "20.00",
            "税率": "6%",
            "税额": "1.20",
        }
    )

    dataframe = build_invoice_dataframe(
        [record],
        mode="items",
        selected_columns=[
            "发票号码",
            "购买方名称",
            "销售方名称",
            "项目名称",
            "明细项目名称",
            "规格型号",
            "税率",
            "明细金额",
            "明细税额",
            "金额合计",
            "税额合计",
            "价税合计大写",
        ],
    )

    assert dataframe.shape == (3, 12)
    assert dataframe.iloc[0]["发票号码"] == "26317000002019064660"
    assert "*服务*第二项目" in dataframe.iloc[0]["项目名称"]
    assert dataframe.iloc[0]["明细项目名称"] == ""
    assert dataframe.iloc[0]["规格型号"] == ""
    assert dataframe.iloc[0]["明细金额"] == ""
    assert dataframe.iloc[0]["明细税额"] == ""
    assert dataframe.iloc[0]["金额合计"] == "55.66"
    assert dataframe.iloc[0]["税额合计"] == "3.34"

    assert dataframe.iloc[1]["发票号码"] == ""
    assert dataframe.iloc[1]["购买方名称"] == ""
    assert dataframe.iloc[1]["销售方名称"] == ""
    assert dataframe.iloc[1]["项目名称"] == ""
    assert dataframe.iloc[1]["明细项目名称"] == "*保险服务*航空旅客人身\n意外伤害保险（互联网）"
    assert dataframe.iloc[1]["规格型号"] == "**"
    assert dataframe.iloc[1]["明细金额"] == "55.66"
    assert dataframe.iloc[1]["明细税额"] == "3.34"
    assert dataframe.iloc[1]["金额合计"] == ""
    assert dataframe.iloc[1]["税额合计"] == ""
    assert dataframe.iloc[1]["价税合计大写"] == ""

    assert dataframe.iloc[2]["发票号码"] == ""
    assert dataframe.iloc[2]["明细项目名称"] == "*服务*第二项目"
    assert dataframe.iloc[2]["明细金额"] == "20.00"
    assert dataframe.iloc[2]["明细税额"] == "1.20"


def test_missing_fields_are_filled_with_blank_strings() -> None:
    dataframe = build_invoice_dataframe([{"文件名称": "missing.pdf", "_status": "ok"}])

    assert dataframe.columns.tolist() == INVOICE_COLUMNS
    assert dataframe.iloc[0]["文件名称"] == "missing.pdf"
    assert dataframe.iloc[0]["发票号码"] == ""


def test_export_xlsx_file(tmp_path: Path) -> None:
    output_path = tmp_path / "invoice.xlsx"

    export_invoices_to_excel(
        [_sample_record()],
        output_path,
        mode="items",
        selected_columns=["文件名称", "发票号码", "项目名称", "税率", "明细金额", "明细税额", "价税合计大写"],
    )

    workbook = load_workbook(output_path)
    sheet = workbook["发票汇总"]

    assert sheet.max_row == 2
    assert sheet.max_column == 7
    assert [cell.value for cell in sheet[1]] == ["文件名称", "发票号码", "项目名称", "税率", "明细金额", "明细税额", "价税合计大写"]
    assert "意外伤害保险（互联网）" in sheet["C2"].value
    assert sheet["C2"].alignment.wrap_text is True
    assert sheet.row_dimensions[2].height == 38
