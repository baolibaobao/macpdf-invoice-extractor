from pathlib import Path

import fitz
from openpyxl import load_workbook

from src.file_manager import BatchProgress, find_pdf_files, process_invoice_folder


def _create_native_text_invoice_pdf(pdf_path: Path) -> None:
    doc = fitz.open()
    page = doc.new_page()
    page.insert_text((72, 72), "Invoice No: 12345678901234567890")
    page.insert_text((72, 92), "This is native text content for batch parser testing.")
    doc.save(pdf_path)
    doc.close()


def test_find_pdf_files_recursively_and_ignore_non_pdf(tmp_path: Path) -> None:
    nested = tmp_path / "nested"
    nested.mkdir()
    (tmp_path / "note.txt").write_text("ignore", encoding="utf-8")
    (tmp_path / "a.PDF").write_bytes(b"%PDF-placeholder")
    (nested / "b.pdf").write_bytes(b"%PDF-placeholder")

    pdf_files = find_pdf_files(tmp_path)

    assert [path.name for path in pdf_files] == ["a.PDF", "b.pdf"]


def test_process_invoice_folder_exports_excel(tmp_path: Path) -> None:
    input_dir = tmp_path / "invoices"
    input_dir.mkdir()
    _create_native_text_invoice_pdf(input_dir / "invoice.pdf")
    (input_dir / "ignore.jpg").write_bytes(b"not a pdf")
    output_path = tmp_path / "out" / "batch.xlsx"
    events: list[BatchProgress] = []

    result = process_invoice_folder(input_dir, output_path, progress_callback=events.append)

    assert result.total_count == 1
    assert result.success_count == 1
    assert result.skipped_count == 0
    assert output_path.exists()
    assert [event.event for event in events] == ["scan", "scan_done", "parse_start", "parse_ok", "export_start", "done"]

    workbook = load_workbook(output_path)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [cell.value for cell in sheet[1]]
    values = [cell.value for cell in sheet[2]]

    assert "发票号码" in headers
    assert values[headers.index("发票号码")] == "12345678901234567890"


def test_process_invoice_folder_keeps_running_when_pdf_is_skipped(tmp_path: Path) -> None:
    input_dir = tmp_path / "invoices"
    input_dir.mkdir()
    empty_pdf = input_dir / "empty.pdf"
    doc = fitz.open()
    doc.new_page()
    doc.save(empty_pdf)
    doc.close()
    output_path = tmp_path / "out.xlsx"

    result = process_invoice_folder(input_dir, output_path)

    assert result.total_count == 1
    assert result.success_count == 0
    assert result.skipped_count == 1
    assert output_path.exists()
